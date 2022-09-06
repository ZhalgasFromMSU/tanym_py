import os
from openpyxl import Workbook

from .models import DatabaseConnector
from .dialogue_texts import PROBLEM_TYPES_STR

def dump_db(db: DatabaseConnector):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Date"
    ws["C1"] = "City"
    ws["D1"] = "Sex"
    ws["E1"] = "Age"
    ws["F1"] = "Type"
    ws["G1"] = "Score"
    ws["H1"] = "Review"

    idx = 2
    for client in db.list_clients():
        ws["A{}".format(idx)] = client.name
        ws["B{}".format(idx)] = client.date.strftime("%d/%m/%Y")
        ws["C{}".format(idx)] = client.city
        ws["D{}".format(idx)] = client.sex
        ws["E{}".format(idx)] = client.age
        ws["F{}".format(idx)] = PROBLEM_TYPES_STR[int(client.pr_type)]
        ws["G{}".format(idx)] = client.score
        ws["H{}".format(idx)] = client.review
        idx += 1

    if os.path.exists("data.xlsx"):
        os.remove("data.xlsx")
    wb.save("data.xlsx")
    return "data.xlsx"
